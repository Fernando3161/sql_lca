"""
Author: Dr.-Ing. Fernando Penaherrera V.
Institution: OFFIS Institue for Computer Sciences
Date: 18.10.2024
License: GNU GENERAL PUBLIC LICENSE
Version: 1.0

This script retrieves and processes data from a database to calculate material contributions for various parts.
The structure of the script is as follows:
1. Access Database
2. Access full Material list
3. Get full inventory for the part
4. Get material demands for each part (unitary)
5. Get partial contributions (material * flow_value)
6. Dump Table into an Excel file
"""

import sqlite3
import pandas as pd
import numpy as np
import os
import sys
import warnings
from openpyxl import Workbook, load_workbook
from sql_connect import create_connection

# Ignore warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=DeprecationWarning)

# Name the Database
parent = os.path.realpath(os.path.join(os.path.abspath('')))
DB_CORR = os.path.join(parent, "databases", "TEMPRO_DB230515_Corr_Rev240119.db")
assert os.path.exists(DB_CORR)

class ExcelFileInput:
    """Class representing input for writing to an Excel file."""
    def __init__(self, df, file_name: str = "Contributions.xlsx", sheet_name: str = "PSID"):
        self.df = df
        self.file_name = file_name
        self.sheet_name = sheet_name

class PSIDContributionInput:
    """Class representing input for calculating PSID contributions."""
    def __init__(self, id: int, file_name: str = "Contributions.xlsx"):
        self.id = id
        self.file_name = file_name

def get_raw_mats_list(db_path=DB_CORR) -> pd.Series:
    """
    Access the database and retrieve the full material list.

    :param db_path: Path to the database file.
    :return: Pandas Series of material names.
    """
    conn = create_connection(db_path)
    query = 'SELECT * FROM [1000Materials]'
    df = pd.read_sql_query(query, conn)
    conn.close()
    return df["Name"]

def get_inventory_by_id(id: int = 3208) -> pd.DataFrame:
    """
    Retrieve full inventory for a specific product ID.

    :param id: Product ID.
    :return: DataFrame containing part exchanges, including amounts.
    """
    conn = create_connection(DB_CORR)
    query = f'SELECT * FROM [3000Exchanges] WHERE "3000ID"={id}'
    df_lci = pd.read_sql_query(query, conn)
    conn.close()

    # Extract relevant columns
    exchanges = df_lci[["2000Parts", "ExchangeName", "Amount"]]
    return exchanges

def get_material_from_flows(exch: pd.DataFrame) -> pd.DataFrame:
    """
    Get the material demand for each part in the exchange list.

    :param exch: DataFrame of part exchanges.
    :return: DataFrame of material flows for each part.
    """
    conn = create_connection(DB_CORR)
    dfs_ = {}
    for part_id in exch["2000Parts"]:
        query = f'SELECT * FROM [2000LCAResults] WHERE "ProductSystemID"={part_id}'
        df_lcia = pd.read_sql_query(query, conn)
        
        # Filter for "EDIP-" categories only
        df_lcia = df_lcia[df_lcia["Category"].str.startswith("EDIP-")]
        df_lcia = df_lcia[["Category", "Result"]]
        df_lcia.set_index("Category", inplace=True)
        dfs_[part_id] = df_lcia
    
    # Concatenate dataframes horizontally
    result = pd.concat(dfs_.values(), axis=1)
    result.columns = dfs_.keys()
    conn.close()
    return result

def get_weighted_contrib(result: pd.DataFrame, exch: pd.DataFrame) -> pd.DataFrame:
    """
    Multiply the value of the impact by the flow value to get the partial contributions.

    :param result: DataFrame of unitary material flows.
    :param exch: DataFrame of part exchanges.
    :return: DataFrame of weighted contributions.
    """
    contrib = result.copy()
    contrib.loc[:, :] = 0  # Initialize all cells to 0

    # Loop over the exchanges to compute weighted contributions
    for _, row in exch.iterrows():
        part_id = row["2000Parts"]
        amount = row["Amount"]

        # Multiply the unitary material demand by the flow value
        for index in contrib.index:
            contrib.at[index, part_id] = result.at[index, part_id] * amount

    # Rename columns from part IDs to exchange names
    contrib.rename(columns={k: v for k, v in zip(exch["2000Parts"], exch["ExchangeName"])}, inplace=True)

    # Fix column names for corrections
    for col in contrib.columns:
        if col.startswith("3") and col[4] == ":":
            contrib.rename(columns={col: "Material Input"}, inplace=True)
    
    return contrib

def fix_index_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove the 'EDIP-' prefix from index names.

    :param df: DataFrame with prefixed index names.
    :return: DataFrame with cleaned index names.
    """
    df.index = df.index.str.replace("^EDIP-", "", regex=True)
    return df

def write_xlsx(input_data: ExcelFileInput) -> None:
    """
    Write DataFrame to an Excel file, overwriting the sheet if it exists.

    :param input_data: ExcelFileInput object containing the DataFrame, file name, and sheet name.
    """
    try:
        # Try to load the existing workbook
        book = load_workbook(input_data.file_name)
        # If the sheet exists, remove it
        if input_data.sheet_name in book.sheetnames:
            std = book[input_data.sheet_name]
            book.remove(std)
    except FileNotFoundError:
        # Create a new workbook if the file does not exist
        book = Workbook()

    # Write the DataFrame to the workbook
    with pd.ExcelWriter(input_data.file_name, engine='openpyxl') as writer:
        writer.book = book
        input_data.df.to_excel(writer, sheet_name=input_data.sheet_name, index=True)
        writer.save()

def get_ps_mat_contribution(input_data: PSIDContributionInput) -> None:
    """
    Calculate material contributions for a specific product system ID (PSID) and save to Excel.

    :param input_data: PSIDContributionInput object containing the PSID and output file name.
    """
    try:
        # Get the exchanges for the specified PSID
        exch = get_inventory_by_id(input_data.id)
        # Retrieve material demands for each part
        unitary_material = get_material_from_flows(exch)
        # Compute weighted contributions
        res = get_weighted_contrib(unitary_material, exch)
        res = fix_index_names(res)
        # Write the result to the Excel file
        write_xlsx(ExcelFileInput(res, input_data.file_name, sheet_name=str(input_data.id)))
        print(f"Finished for PS={input_data.id}")
    except Exception as e:
        print(f"Not possible for PS={input_data.id}", "--", e)

def main() -> None:
    """
    Main function to process multiple product system IDs and calculate their material contributions.
    """
    for psid in range(3200, 3230):
        get_ps_mat_contribution(PSIDContributionInput(id=psid))

if __name__ == "__main__":
    main()
